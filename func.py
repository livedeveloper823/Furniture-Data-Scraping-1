from selenium import webdriver
from selenium.webdriver.remote.webelement import WebElement
from time import sleep
import openpyxl

def wait_url(driver : webdriver.Chrome, url : str):
    print(url)
    while True:
        cur_url = driver.current_url
        if cur_url == url:
            break
        sleep(0.1)  

def find_element(driver : webdriver.Chrome, whichBy, unique : str) -> WebElement:
    while True:
        try:
            element = driver.find_element(whichBy, unique)
            break
        except:
            pass
        sleep(1)
    return element

def find_elements(driver : webdriver.Chrome, whichBy, unique : str) -> list[WebElement]:
    while True:
        try:
            elements = driver.find_elements(whichBy, unique)
            break
        except:
            pass
        sleep(0.1)
    return elements


def create_excel():
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet.merge_cells('F1:M1')
    worksheet.merge_cells('N1:AU1')

    worksheet['F1'] = "Dimensiones"
    worksheet['N1'] = "Especificaciones"

    worksheet['A1'] = "Categoría"
    worksheet['B1'] = "Línea"
    worksheet['C1'] = "Sublinea"
    worksheet['D2'] = "Nombre"
    worksheet['E2'] = "Precio"
    worksheet['F2'] = "Ancho"
    worksheet['G2'] = "Alto"
    worksheet['H2'] = "Peso"
    worksheet['I2'] = "Fondo"
    worksheet['J2'] = "Largo"
    worksheet['K2'] = "Profundidad"
    worksheet['L2'] = "Peso máximo soportado"
    worksheet['M2'] = "Dimensiones de la cama"
    worksheet['N2'] = "Número de cuerpos"
    worksheet['O2'] = "Material del tapiz"
    worksheet['P2'] = "Términos Garantía"
    worksheet['Q2'] = "País de Origen"
    worksheet['R2'] = "Material de la estructura"
    worksheet['S2'] = "Garantía Producto"
    worksheet['T2'] = "Tipo"
    worksheet['U2'] = "Color del tapiz"
    worksheet['V2'] = "Material"
    worksheet['W2'] = "Color"
    worksheet['X2'] = "Número de puestos"
    worksheet['Y2'] = "Pulgadas TV"
    worksheet['Z2'] = "Diseño de la Mesa"
    worksheet['AA2'] = "Plegable"
    worksheet['AB2'] = "Cuenta con ruedas"
    worksheet['AC2'] = "Modo de fijación"
    worksheet['AD2'] = "Espacio recomendado"
    worksheet['AE2'] = "Cantidad de repisas"
    worksheet['AF2'] = "Cantidad de cajones"
    worksheet['AG2'] = "Cantidad de puertas"
    worksheet['AH2'] = "Tipo de guadarropa"
    worksheet['AI2'] = "Sistema de apertura"
    worksheet['AJ2'] = "Número de cajones"
    worksheet['AK2'] = "Tamaño"
    worksheet['AL2'] = "Nivel de Confort"
    worksheet['AM2'] = "Composición Interna"
    worksheet['AN2'] = "Línea"
    worksheet['AO2'] = "Color del Colchón"
    worksheet['AP2'] = "Img_url"

    workbook.save('homecenter.xlsx')